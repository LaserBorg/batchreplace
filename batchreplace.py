"""
  _           _       _                    _
 | |__   __ _| |_ ___| |__  _ __ ___ _ __ | | __ _  ___ ___
 | '_ \ / _` | __/ __| '_ \| '__/ _ \ '_ \| |/ _` |/ __/ _ \
 | |_) | (_| | || (__| | | | | |  __/ |_) | | (_| | (_|  __/
 |_.__/ \__,_|\__\___|_| |_|_|  \___| .__/|_|\__,_|\___\___|
                                    |_|
"""
import re
import os
from openpyxl import load_workbook


# convert xlsx into *list of rows*
def read_xlsx(filepath, hasHeader=True):
    ws = load_workbook(filepath).active

    rows = []
    for ws_row in ws.iter_rows(min_row=1 + hasHeader):
        rows.append([cell.value for cell in ws_row])

    header = [cell.value for cell in ws[1]] if hasHeader else None
    return rows, header


# remove irrelevant text, just keep the url
def clean_rows(raw_rows, column=2, left='src="', right='"'):
    rows = []
    for row in raw_rows:
        src = select_str_by_delimiters(row[column], left=left, right=right)
        # skip rows without url
        if src is not None:
            rows.append([row[0], row[1], src])
    return rows


# use regex to find values
def select_str_by_delimiters(text, left='src="', right='"'):
    result = re.search(f'{left}(.*){right}', text)
    if result is not None:
        result = result.group(1)
    return result


content_path = "input/content.xlsx"
template_path = 'input/template.html'
output_dir = 'output'

col_branch =   0
col_filename = 1
col_url =      2
ext = os.path.splitext(template_path)[1]

# create output directory if necessary
if not os.path.exists(output_dir):
    os.makedirs(output_dir)


# read xlsx
raw_rows, header = read_xlsx(content_path)
rows = clean_rows(raw_rows, column=col_url, left='src="', right='"')

# read template file
with open(template_path, 'r') as file:
    template_data = file.read()

# iterate through all rows of cleaned data
print("")
print(header)
for row in rows:
    print(row)

    data = template_data
    # replace all variables
    data = data.replace("$TITLE", row[col_filename])
    data = data.replace("$SOURCE", row[col_url])

    # write output file
    output_file = f"{row[col_branch]}_{row[col_filename]}{ext}"
    with open(os.path.join(output_dir, output_file), 'w') as file:
        file.write(data)

print("finished.")
